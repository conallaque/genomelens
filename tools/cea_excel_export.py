#!/usr/bin/env python3
"""Export the Markov cost-effectiveness model as a LIVE Excel workbook.

Why this exists
---------------
`econ/markov.py` computes the model in Python and the report prints the answer.
That is a *report*. What an HEOR analyst hands a payer is a *model*: a workbook
whose inputs are editable cells and whose ICER is a formula, so the reader can
change the discount rate and watch the answer move. This writes that workbook.

Every number in `Trace_*` and `Results` is an Excel formula referencing named
cells on `Inputs`. Nothing is a baked-in Python result, which is the whole
point — a static export would be indistinguishable from a screenshot.

The workbook reproduces `markov.run_markov` / `markov.markov_cost_effectiveness`
exactly, including the details that are easy to get wrong and that an HTA
reviewer checks:

* rate → probability via ``p = 1 - exp(-rate * cycle_length)``, not ``p = rate``
* competing risks from Well, normalized so the two exits cannot exceed 1
* ``p_dd`` built from the ORIGINAL background rate, not the normalized one
  (mirrors the local-rebinding subtlety in ``build_transition_matrix``)
* half-cycle correction weighting cycle 0 and the final cycle at 0.5
* age-dependent Gompertz background mortality
* ICER suppressed in the dominance quadrants, where the ratio is ambiguous

Standalone by design: no pipeline wiring, no CLI flag, no new runtime
dependency for the main package. openpyxl only, and only when you run this.

Usage
-----
    python tools/cea_excel_export.py                    # -> cea_model.xlsx
    python tools/cea_excel_export.py --out /tmp/m.xlsx --cycles 30
    python tools/cea_excel_export.py --fixture           # also write the R fixture
"""

from __future__ import annotations

import argparse
import csv
import json
import sys
from pathlib import Path

_REPO = Path(__file__).resolve().parent.parent
if str(_REPO) not in sys.path:
    sys.path.insert(0, str(_REPO))

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.workbook.defined_name import DefinedName
except ImportError:  # pragma: no cover - dependency is optional by design
    sys.exit("openpyxl is required: pip install openpyxl")


# ── model inputs ──────────────────────────────────────────────────────────────
# Name, default, number format, unit, note. Order is the sheet order, and the
# name is the Excel defined-name every formula below refers to. Defaults match
# econ/markov.run_markov so a fresh workbook reproduces the Python answer.
INPUTS: list[tuple[str, float, str, str, str]] = [
    ("start_age", 40.0, "0", "years", "Cohort age at cycle 0"),
    ("cycles", 45.0, "0", "cycles", "Annual cycles. Lowering this works; raising it needs a re-export."),
    ("incidence_rate", 0.010, "0.0000", "per year", "Well -> Disease hazard RATE (not a probability)"),
    ("rrr_intervention", 0.30, "0.0%", "fraction", "Relative risk reduction applied to incidence in the guided arm"),
    ("cost_intervention_annual", 250.0, '"$"#,##0', "$/yr", "Paid while the patient remains Well, guided arm only"),
    ("cost_disease_annual", 12000.0, '"$"#,##0', "$/yr", "Annual cost in the Disease state"),
    ("cost_well_annual", 0.0, '"$"#,##0', "$/yr", "Annual cost in the Well state, both arms"),
    ("utility_well", 0.90, "0.00", "QALY/yr", "Health-related quality of life, Well"),
    ("utility_disease", 0.68, "0.00", "QALY/yr", "Health-related quality of life, Disease"),
    ("excess_mortality_rate", 0.035, "0.000", "per year", "Disease-attributable excess mortality RATE"),
    ("discount_rate", 0.03, "0.0%", "per year", "Applied to both costs and QALYs"),
    ("wtp", 100000.0, '"$"#,##0', "$/QALY", "Cost-effectiveness threshold"),
    ("gompertz_a", 0.0001, "0.00000", "-", "Gompertz scale for background mortality"),
    ("gompertz_b", 0.085, "0.000", "-", "Gompertz shape for background mortality"),
    ("cycle_length", 1.0, "0.0", "years", "Cycle length used in the rate->probability conversion"),
    ("half_cycle", 1.0, "0", "1=on 0=off", "Weight cycle 0 and the final cycle at 0.5"),
]

# Trace columns: header, width, number format. Column letters are derived from
# position, so inserting a column here updates every formula that uses _c().
TRACE_COLS: list[tuple[str, int, str]] = [
    ("cycle", 7, "0"),
    ("age", 7, "0.0"),
    ("incidence (eff. rate)", 20, "0.00000"),
    ("p_wd raw", 11, "0.00000"),
    ("p_bg raw", 11, "0.00000"),
    ("p_ex", 11, "0.00000"),
    ("p_wd norm", 11, "0.00000"),
    ("p_bg norm", 11, "0.00000"),
    ("p_dd", 11, "0.00000"),
    ("Well", 11, "0.00000"),
    ("Disease", 11, "0.00000"),
    ("Dead", 11, "0.00000"),
    ("cost/cycle", 13, '"$"#,##0.00'),
    ("qaly/cycle", 12, "0.00000"),
    ("ly/cycle", 11, "0.00000"),
    ("weight", 9, "0.0"),
    ("discount", 11, "0.00000"),
    ("disc. cost", 14, '"$"#,##0.00'),
    ("disc. qaly", 12, "0.00000"),
    ("disc. ly", 11, "0.00000"),
]

_COL = {name: get_column_letter(i + 1) for i, (name, _, _) in enumerate(TRACE_COLS)}

HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(color="FFFFFF", bold=True, size=10)
IN_FILL = PatternFill("solid", fgColor="FFF2CC")   # editable cells
OUT_FILL = PatternFill("solid", fgColor="E2EFDA")  # computed headline cells
TITLE_FONT = Font(bold=True, size=13, color="1F3864")
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _c(col: str, row: int) -> str:
    """Trace-sheet cell reference by column NAME, so formulas survive reordering."""
    return f"{_COL[col]}{row}"


def _build_inputs(ws) -> None:
    ws["A1"] = "GenomeLens - Markov cost-effectiveness model"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = ("Yellow cells are inputs: edit them and every figure on Results "
                "recalculates. Nothing here is a pasted Python result.")
    ws["A2"].font = Font(italic=True, size=9, color="808080")

    for col, hdr, width in (("A", "Parameter", 26), ("B", "Value", 14),
                            ("C", "Units", 12), ("D", "Note", 74)):
        ws[f"{col}4"] = hdr
        ws[f"{col}4"].fill = HDR_FILL
        ws[f"{col}4"].font = HDR_FONT
        ws.column_dimensions[col].width = width

    for i, (name, default, fmt, unit, note) in enumerate(INPUTS):
        r = 5 + i
        ws[f"A{r}"] = name
        ws[f"A{r}"].font = Font(bold=True, size=10)
        ws[f"B{r}"] = default
        ws[f"B{r}"].number_format = fmt
        ws[f"B{r}"].fill = IN_FILL
        ws[f"B{r}"].border = BOX
        ws[f"C{r}"] = unit
        ws[f"D{r}"] = note
        ws[f"D{r}"].alignment = Alignment(wrap_text=False)
        ws[f"D{r}"].font = Font(size=9, color="595959")


def _build_trace(ws, *, guided: bool, cycles: int) -> None:
    """One arm's cohort trace. Every cell is a formula except cycle index and seed."""
    arm = "Genomic-guided" if guided else "Standard care"
    ws["A1"] = f"{arm} - cohort trace (Well / Disease / Dead)"
    ws["A1"].font = TITLE_FONT

    for i, (hdr, width, _) in enumerate(TRACE_COLS):
        cell = ws.cell(row=3, column=i + 1, value=hdr)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        ws.column_dimensions[get_column_letter(i + 1)].width = width
    ws.freeze_panes = "A4"

    first = 4
    for t in range(cycles):
        r = first + t
        p = r - 1  # previous row

        ws[_c("cycle", r)] = t
        ws[_c("age", r)] = f"=start_age+{t}"

        # Effective incidence: the guided arm applies the RRR to the hazard.
        ws[_c("incidence (eff. rate)", r)] = (
            "=incidence_rate*(1-rrr_intervention)" if guided else "=incidence_rate")

        # Rate -> probability. This is the conversion HTA reviewers check for.
        ws[_c("p_wd raw", r)] = f"=1-EXP(-{_c('incidence (eff. rate)', r)}*cycle_length)"
        ws[_c("p_bg raw", r)] = (
            f"=1-EXP(-(gompertz_a*EXP(gompertz_b*MAX(0,{_c('age', r)}-20)))*cycle_length)")
        ws[_c("p_ex", r)] = "=1-EXP(-excess_mortality_rate*cycle_length)"

        # Competing risks from Well: if the two exits sum above 1, scale BOTH.
        tot = f"({_c('p_wd raw', r)}+{_c('p_bg raw', r)})"
        ws[_c("p_wd norm", r)] = (
            f"=IF({tot}>1,{_c('p_wd raw', r)}/{tot},{_c('p_wd raw', r)})")
        ws[_c("p_bg norm", r)] = (
            f"=IF({tot}>1,{_c('p_bg raw', r)}/{tot},{_c('p_bg raw', r)})")

        # p_dd uses the ORIGINAL background probability, not the normalized one.
        # build_transition_matrix rebinds its local p_bg before this line but
        # reads the parameter, so the un-normalized value is what reaches here.
        ws[_c("p_dd", r)] = f"=MIN(1,{_c('p_bg raw', r)}+{_c('p_ex', r)})"

        # Cohort occupancy at the START of the cycle.
        if t == 0:
            ws[_c("Well", r)] = 1.0
            ws[_c("Disease", r)] = 0.0
            ws[_c("Dead", r)] = 0.0
        else:
            ws[_c("Well", r)] = (
                f"={_c('Well', p)}*(1-{_c('p_wd norm', p)}-{_c('p_bg norm', p)})")
            ws[_c("Disease", r)] = (
                f"={_c('Well', p)}*{_c('p_wd norm', p)}"
                f"+{_c('Disease', p)}*(1-{_c('p_dd', p)})")
            ws[_c("Dead", r)] = (
                f"={_c('Dead', p)}+{_c('Well', p)}*{_c('p_bg norm', p)}"
                f"+{_c('Disease', p)}*{_c('p_dd', p)}")

        # Payoffs valued on start-of-cycle occupancy.
        interv = "+cost_intervention_annual" if guided else ""
        ws[_c("cost/cycle", r)] = (
            f"={_c('Well', r)}*(cost_well_annual{interv})"
            f"+{_c('Disease', r)}*cost_disease_annual")
        ws[_c("qaly/cycle", r)] = (
            f"={_c('Well', r)}*utility_well+{_c('Disease', r)}*utility_disease")
        ws[_c("ly/cycle", r)] = f"={_c('Well', r)}+{_c('Disease', r)}"

        # Half-cycle correction, driven by the live `cycles` input so shrinking
        # the horizon moves the final-cycle weight to the right row.
        ws[_c("weight", r)] = (
            f"=IF({_c('cycle', r)}>cycles-1,0,"
            f"IF(AND(half_cycle=1,OR({_c('cycle', r)}=0,{_c('cycle', r)}=cycles-1)),0.5,1))")
        ws[_c("discount", r)] = f"=1/(1+discount_rate)^{_c('cycle', r)}"

        for src, dst in (("cost/cycle", "disc. cost"), ("qaly/cycle", "disc. qaly"),
                         ("ly/cycle", "disc. ly")):
            ws[_c(dst, r)] = (
                f"={_c(src, r)}*{_c('weight', r)}*{_c('discount', r)}")

        for name, _, fmt in TRACE_COLS:
            ws[_c(name, r)].number_format = fmt

    # Totals row.
    last = first + cycles - 1
    tr = last + 2
    ws[f"A{tr}"] = "TOTAL (discounted)"
    ws[f"A{tr}"].font = Font(bold=True)
    for name in ("disc. cost", "disc. qaly", "disc. ly"):
        col = _COL[name]
        cell = ws[f"{col}{tr}"]
        cell.value = f"=SUM({col}{first}:{col}{last})"
        cell.number_format = dict((n, f) for n, _, f in TRACE_COLS)[name]
        cell.font = Font(bold=True)
        cell.fill = OUT_FILL
        cell.border = BOX
    return None


def _totals_ref(sheet: str, cycles: int, name: str) -> str:
    return f"'{sheet}'!{_COL[name]}{4 + cycles + 1}"


def _build_results(ws, *, cycles: int) -> None:
    ws["A1"] = "Incremental cost-effectiveness"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = ("Genomic-guided vs standard care. Every cell below is a formula "
                "over the two trace sheets; change any input and this moves.")
    ws["A2"].font = Font(italic=True, size=9, color="808080")

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 76

    sc_c = _totals_ref("Trace_SC", cycles, "disc. cost")
    sc_q = _totals_ref("Trace_SC", cycles, "disc. qaly")
    sc_l = _totals_ref("Trace_SC", cycles, "disc. ly")
    gg_c = _totals_ref("Trace_GG", cycles, "disc. cost")
    gg_q = _totals_ref("Trace_GG", cycles, "disc. qaly")
    gg_l = _totals_ref("Trace_GG", cycles, "disc. ly")

    rows: list[tuple[str, str, str, str, bool]] = [
        ("Standard care - total cost", f"={sc_c}", '"$"#,##0', "Discounted, half-cycle corrected", False),
        ("Standard care - total QALYs", f"={sc_q}", "0.0000", "", False),
        ("Standard care - life years", f"={sc_l}", "0.0000", "", False),
        ("Genomic-guided - total cost", f"={gg_c}", '"$"#,##0', "Includes the annual intervention cost while Well", False),
        ("Genomic-guided - total QALYs", f"={gg_q}", "0.0000", "", False),
        ("Genomic-guided - life years", f"={gg_l}", "0.0000", "", False),
        ("Incremental cost (dC)", f"={gg_c}-{sc_c}", '"$"#,##0', "Guided minus standard care", True),
        ("Incremental QALYs (dQ)", f"={gg_q}-{sc_q}", "0.0000", "", True),
        ("Incremental life years", f"={gg_l}-{sc_l}", "0.0000", "", False),
    ]
    r = 4
    for label, formula, fmt, note, headline in rows:
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(bold=headline, size=10)
        c = ws[f"B{r}"]
        c.value = formula
        c.number_format = fmt
        c.border = BOX
        if headline:
            c.fill = OUT_FILL
            c.font = Font(bold=True)
        ws[f"C{r}"] = note
        ws[f"C{r}"].font = Font(size=9, color="595959")
        r += 1

    dc, dq = "B10", "B11"   # incremental cost / QALY cells

    r += 1
    ws[f"A{r}"] = "Dominant (cheaper AND more effective)"
    ws[f"B{r}"] = f"=IF(AND({dc}<0,{dq}>0),\"YES\",\"no\")"
    dom_row = r
    r += 1
    ws[f"A{r}"] = "Dominated (costlier AND less effective)"
    ws[f"B{r}"] = f"=IF(AND({dc}>0,{dq}<0),\"YES\",\"no\")"
    dmd_row = r
    r += 1

    ws[f"A{r}"] = "ICER ($/QALY)"
    ws[f"A{r}"].font = Font(bold=True)
    # An ICER is only interpretable when cost and effect move the same way. In
    # the dominance quadrants HTA convention is to state dominance and withhold
    # the ratio, because "-$6,054/QALY" reads as a bargain either way.
    icer = ws[f"B{r}"]
    icer.value = (f'=IF(OR(B{dom_row}="YES",B{dmd_row}="YES"),"suppressed (dominance)",'
                  f'IF(ABS({dq})<1E-09,"undefined",{dc}/{dq}))')
    icer.number_format = '"$"#,##0'
    icer.fill = OUT_FILL
    icer.font = Font(bold=True)
    icer.border = BOX
    ws[f"C{r}"] = ("Suppressed in the dominance quadrants by HTA reporting "
                   "convention, matching markov_cost_effectiveness.")
    ws[f"C{r}"].font = Font(size=9, color="595959")
    r += 1

    ws[f"A{r}"] = "NMB at WTP"
    ws[f"A{r}"].font = Font(bold=True)
    nmb = ws[f"B{r}"]
    nmb.value = f"=wtp*{dq}-{dc}"
    nmb.number_format = '"$"#,##0'
    nmb.fill = OUT_FILL
    nmb.font = Font(bold=True)
    nmb.border = BOX
    ws[f"C{r}"] = "Net monetary benefit. Positive = worth doing at the threshold."
    ws[f"C{r}"].font = Font(size=9, color="595959")
    nmb_row = r
    r += 2

    ws[f"A{r}"] = "Verdict"
    ws[f"A{r}"].font = Font(bold=True)
    v = ws[f"B{r}"]
    v.value = (f'=IF(B{dom_row}="YES","dominant (cheaper and more effective)",'
               f'IF(B{dmd_row}="YES","dominated (costlier and less effective)",'
               f'IF(B{nmb_row}>=0,"cost-effective at WTP","not cost-effective at WTP")))')
    v.font = Font(bold=True)
    v.fill = OUT_FILL
    v.border = BOX
    ws.merge_cells(f"C{r}:C{r}")
    r += 2

    ws[f"A{r}"] = "Model note"
    ws[f"A{r}"].font = Font(bold=True)
    ws[f"B{r}"] = ("Discrete-time Markov cohort model, annual cycles, half-cycle "
                   "correction, p = 1 - exp(-rate*dt), age-dependent Gompertz "
                   "competing mortality, 3% discounting of costs and QALYs. "
                   "Mirrors econ/markov.py. Sonnenberg & Beck (1993); Briggs, "
                   "Sculpher & Claxton (2006).")
    ws[f"B{r}"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[r].height = 46
    ws.merge_cells(f"B{r}:C{r}")


def build_workbook(out_path: Path, *, cycles: int = 45) -> Path:
    wb = Workbook()
    inputs = wb.active
    inputs.title = "Inputs"
    _build_inputs(inputs)

    for i, (name, *_rest) in enumerate(INPUTS):
        wb.defined_names.add(DefinedName(name, attr_text=f"Inputs!$B${5 + i}"))

    _build_trace(wb.create_sheet("Trace_SC"), guided=False, cycles=cycles)
    _build_trace(wb.create_sheet("Trace_GG"), guided=True, cycles=cycles)
    _build_results(wb.create_sheet("Results"), cycles=cycles)

    # Results first: it is the answer, and a workbook that opens on a 45-row
    # trace looks like a spreadsheet rather than a model.
    wb.move_sheet("Results", offset=-3)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def python_reference(cycles: int = 45) -> dict:
    """Python's answer for the workbook defaults, for cross-validation.

    Parameter-driven: this touches no genome, no report artifact and no
    personal data. It calls the econ engine with the same defaults the
    workbook ships with.
    """
    from econ import markov as mk

    defaults = {k: v for k, v, *_ in INPUTS}
    res = mk.markov_cost_effectiveness(
        wtp=defaults["wtp"],
        start_age=defaults["start_age"],
        cycles=cycles,
        incidence_rate=defaults["incidence_rate"],
        rrr_intervention=defaults["rrr_intervention"],
        cost_intervention_annual=defaults["cost_intervention_annual"],
        cost_disease_annual=defaults["cost_disease_annual"],
        cost_well_annual=defaults["cost_well_annual"],
        utility_well=defaults["utility_well"],
        utility_disease=defaults["utility_disease"],
        excess_mortality_rate=defaults["excess_mortality_rate"],
        discount_rate=defaults["discount_rate"],
        half_cycle=bool(defaults["half_cycle"]),
    )
    return {
        "inputs": {**defaults, "cycles": float(cycles)},
        "expected": {
            "sc_cost": res["standard_care"]["total_cost"],
            "sc_qaly": res["standard_care"]["total_qaly"],
            "sc_ly": res["standard_care"]["total_life_years"],
            "gg_cost": res["genomic_guided"]["total_cost"],
            "gg_qaly": res["genomic_guided"]["total_qaly"],
            "gg_ly": res["genomic_guided"]["total_life_years"],
            "incremental_cost": res["incremental_cost"],
            "incremental_qaly": res["incremental_qaly"],
            "icer": res["icer"],
            "nmb_at_wtp": res["nmb_at_wtp"],
            "dominant": res["dominant"],
            "dominated": res["dominated"],
            "verdict": res["verdict"],
        },
        "note": ("Generated by tools/cea_excel_export.py from econ/markov.py. "
                 "Synthetic parameters only - no genome or personal data."),
    }


def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser(description=__doc__.split("\n")[0])
    # Default into build/, which .gitignore already covers. The repo's other
    # generated artefacts land in the root and are ignored BY NAME, because the
    # pipeline has to write them beside the genome it analyzed and cannot choose
    # a directory. This tool can choose. Defaulting to the root would add
    # another filename to that list -- and the list has already failed once
    # (see the economics.html note in .gitignore, committed by accident because
    # the name was new and no rule covered it). An already-ignored directory
    # needs no new rule and is safe on a fresh clone.
    ap.add_argument("--out", default=str(_REPO / "build" / "cea_model.xlsx"),
                    help="Output .xlsx path (default: build/cea_model.xlsx)")
    ap.add_argument("--cycles", type=int, default=45,
                    help="Annual cycles to generate (default 45)")
    ap.add_argument("--fixture", action="store_true",
                    help="Also write tools/cea_fixture.json for the R check")
    # Writing the committed sample is an explicit act, never the default —
    # same convention as scripts/make_econ_sample.py, which names its
    # docs/samples/ destinations outright. The default output stays in
    # build/ so a routine run cannot land a stale binary in a tracked path.
    ap.add_argument("--sample", action="store_true",
                    help=("Write the README's committed sample to "
                          "docs/samples/cea-model-sample.xlsx (overrides --out)."))
    a = ap.parse_args(argv)

    out = (_REPO / "docs" / "samples" / "cea-model-sample.xlsx") if a.sample \
        else Path(a.out)
    path = build_workbook(out, cycles=a.cycles)
    print(f"Excel model written: {path}")
    print("  Inputs (yellow) are editable; Results recalculates from formulas.")

    if a.fixture:
        here = Path(__file__).resolve().parent
        ref = python_reference(cycles=a.cycles)
        fx = here / "cea_fixture.json"
        fx.write_text(json.dumps(ref, indent=2) + "\n")
        print(f"R cross-validation fixture written: {fx}")

        # Flat CSV as well: markov_check.R is deliberately base-R only, and base
        # R has no JSON parser. read.csv does this in one line with no packages.
        # TAB-separated, not comma. The repo's .gitignore blanket-ignores *.csv
        # as part of the never-commit-real-DNA rule, with path-anchored
        # exceptions only. A .csv fixture here would be silently untracked and
        # the R check would fail on a fresh clone. .tsv sidesteps that without
        # touching the DNA-safety gitignore, and base R reads it with read.delim.
        tsv_path = here / "cea_fixture.tsv"

        def _tok(v: object) -> str:
            # Bare tokens R can coerce: numbers as-is, None -> NA, bools as
            # R literals, strings plain.
            if v is None:
                return "NA"
            if isinstance(v, bool):
                return "TRUE" if v else "FALSE"
            return str(v)

        with tsv_path.open("w", newline="") as fh:
            w = csv.writer(fh, delimiter="\t")
            w.writerow(["section", "key", "value"])
            for k, v in ref["inputs"].items():
                w.writerow(["input", k, _tok(v)])
            for k, v in ref["expected"].items():
                w.writerow(["expected", k, _tok(v)])
        print(f"                              and: {tsv_path}")

        e = ref["expected"]
        print(f"  Python: dC={e['incremental_cost']:,.2f}  "
              f"dQ={e['incremental_qaly']:.4f}  NMB={e['nmb_at_wtp']:,.2f}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
